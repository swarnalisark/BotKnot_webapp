from flask import Flask, session, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import pickle
from datetime import timedelta
from scipy.ndimage import gaussian_filter1d
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from io import BytesIO
import base64

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Needed for flash messages

# Define Excel file for storing user data
EXCEL_FILE = 'user_data.xlsx'
PRODUCT_EXCEL_FILE = 'Web_Scraping.xlsx'

ITEM_CODE_MAPPING = {
    "Allgäuer Hof-Milch Butter mild gesäuert 250g": "101",
    "Bio Aubergine 1 Stück": "201",
    "Blumenkohl weiß 1 Stück": "301",
    "Broccoli 500g": "401",
    "Eisbergsalat 1 Stück": "501",
    "Galiamelone 1 Stück": "601",
    "Karotten 1kg": "701",
    "Kartoffeln vorwiegend festkochend 2,5kg": "801",
    "Mango vorgereift 1 Stück": "901",
    "Meggle Feine Butter 250g": "1001",
    "Orangen 2kg im Netz": "1101",
    "REWE Beste Wahl Feinschmecker Hähnchen 1200g": "1201",
    "REWE Bio Zucchini 500g": "1301",
    "Rewe Beste Wahl Eier aus Freilandhaltung 10 Stück": "1401",
    "Rispentomaten ca. 100g": "1501",
    "Spitzkohl ca. 1kg": "1601",
    "Tafeltrauben hell kernlos 500g": "1701",
    "Zitronen 500g im Netz": "1801",
    "Zwiebeln 2kg im Netz": "1901",
    "ja! Basmati Reis 1kg": "2001",
    "ja! H-Milch 3,5% 1": "2101",
    "ja! Sonnenblumenöl 1l": "2201"
}

ITEM_CODE_TO_ENGLISH_NAME = {
    "101": "Butter Mildly Soured",
    "201": "Eggplant",
    "301": "White Cauliflower",
    "401": "Broccoli",
    "501": "Iceberg Lettuce",
    "601": "Galia Melon",
    "701": "Carrots",
    "801": "Potatoes",
    "901": "Mango",
    "1001": "Meggle Fine Butter",
    "1101": "Oranges",
    "1201": "Chicken",
    "1301": "Zucchini",
    "1401": "Eggs",
    "1501": "Tomatoes",
    "1601": "Cabbage",
    "1701": "Table Grapes",
    "1801": "Lemons",
    "1901": "Onions",
    "2001": "Basmati Rice",
    "2101": "Milk 3.5%",
    "2201": "Sunflower Oil"
}

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        ws.append(['UserID', 'Password'])  # Headers
        wb.save(EXCEL_FILE)

def load_products():
    try:
        wb = load_workbook(PRODUCT_EXCEL_FILE)
        ws = wb.active

        # Add new columns if they do not exist
        headers = [cell.value for cell in ws[1]]  # Get existing headers

        if "Item Code" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="Item Code")
            headers.append("Item Code")
        if "English Name" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="English Name")
            headers.append("English Name")

        product_name_idx = headers.index("Product Name") + 1
        item_code_idx = headers.index("Item Code") + 1
        english_name_idx = headers.index("English Name") + 1

        # Update the Excel file with mappings
        for row in ws.iter_rows(min_row=2, max_col=len(headers)):
            product_name = row[product_name_idx - 1].value
            if product_name in ITEM_CODE_MAPPING:
                item_code = ITEM_CODE_MAPPING[product_name]
                row[item_code_idx - 1].value = item_code
                english_name = ITEM_CODE_TO_ENGLISH_NAME.get(item_code, "Unknown")
                row[english_name_idx - 1].value = english_name

        wb.save(PRODUCT_EXCEL_FILE)  # Save updated file

        # Load products into a list
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(row):
                product = {
                    'name': row[product_name_idx - 1],
                    'price': row[1],
                    'tomorrow_price': row[2],
                    'image': row[3],
                    'item_code': row[item_code_idx - 1],
                    'english_name': row[english_name_idx - 1]
                }
                products.append(product)

        return products
    except Exception as e:
        print("Error loading products:", e)
        return []

def predict_and_save(item_name, models, data, steps=90, output_folder="Datasets/Output/Price_Prediction", csv_file="Datasets/Output/Price_Prediction/Item_lists.csv"):
    if item_name not in models:
        print(f"No model found for item: {item_name}")
        return None, None

    model = models[item_name]
    item_data = data[data['Items'] == item_name]
    item_data['Date'] = pd.to_datetime(item_data['Date'])  # Ensure 'Date' is datetime
    daily_prices = item_data.set_index('Date')['price']

    # Forecast future values
    forecast = model.forecast(steps=steps)
    last_date = daily_prices.index[-1]
    future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=steps)

    # Smooth historical prices
    smoothed_prices = gaussian_filter1d(daily_prices, sigma=2)

    # Smooth forecasted prices
    smoothed_forecast = gaussian_filter1d(forecast, sigma=2)

    # Save forecast to summary data
    price_for_tomorrow = forecast.iloc[0]
    price_after_7_days = forecast.iloc[7] if len(forecast) > 7 else None
    price_after_1_month = forecast.iloc[30] if len(forecast) > 30 else None

    item_code = ITEM_CODE_MAPPING.get(item_name, "Unknown")
    item_english_name = ITEM_CODE_TO_ENGLISH_NAME.get(item_code, "Unknown")

    summary_data = {
        "Serial_Number": len(pd.read_csv(csv_file)) + 1 if os.path.exists(csv_file) else 1,
        "Item_Code": item_code,
        "Item_Name": item_name,
        "Price_for_Tomorrow": price_for_tomorrow,
        "Price_After_7_Days": price_after_7_days,
        "Price_After_1_Month": price_after_1_month
    }

    # Plot
    plt.figure(figsize=(12, 6))
    plt.plot(daily_prices.index, smoothed_prices, label="Actual Price", linewidth=2, linestyle='-', color="blue")
    plt.plot(future_dates, smoothed_forecast, label="Predicted Price (Next 90 Days)", linewidth=3, linestyle='-', color='orange')
    plt.title(f"Price Trend for {item_english_name}", fontsize=16)
    plt.ylabel("Price (€)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    # Save plot to a BytesIO object
    buf = BytesIO()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plot_url = base64.b64encode(buf.getvalue()).decode("utf-8")
    buf.close()
    plt.close()

    return summary_data, plot_url

@app.route('/product/<product_name>')
def product_details(product_name):
    grocery_data = pd.read_csv('Datasets/Price_Predictions/grocery_items_bavaria.csv')
    grocery_data.rename(columns={'name': 'Items'}, inplace=True)

    models_folder = "Model/Price_Prediction/arima"
    with open(os.path.join(models_folder, "arima_models.pkl"), 'rb') as f:
        loaded_models = pickle.load(f)

    summary_data, plot_url = predict_and_save(product_name, loaded_models, grocery_data)

    if summary_data is None or plot_url is None:
        flash("Unable to fetch product details.", "error")
        return redirect(url_for("dashboard", user_id=session.get("user_id")))

    return render_template("product_details.html", product_name=product_name, summary_data=summary_data, plot_url=plot_url)

@app.before_request
def initialize_cart():
    if 'cart' not in session:
        session['cart'] = []  # Initialize cart as an empty list

@app.route('/')
def home():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    user_id = request.form['user_id']
    password = request.form['password']

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    user_found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_found = True
            if row[1] == password:
                 # Store user details in session
                session['user_id'] = user_id
                return redirect(url_for('dashboard', user_id=user_id))
            else:
                flash('Wrong password. Please try again.', 'error')
                return redirect(url_for('home'))

    if not user_found:
        flash('User ID not found. Please sign up if you are not registered.', 'error')
    return redirect(url_for('home'))

@app.route('/dashboard/<user_id>')
def dashboard(user_id):
    search_query = request.args.get('search_query', '').strip().lower()
    selected_product = request.args.get('product_name', '').strip()

    # Load products
    products = load_products()

    # Filter products if search query is provided
    if search_query:
        products = [product for product in products if search_query in product['name'].lower()]
        if not products:
            flash('Product not available.', 'error')

    # Prepare data for the selected product
    summary_data = None
    plot_url = None
    if selected_product:
        grocery_data = pd.read_csv('Datasets/Price_Predictions/grocery_items_bavaria.csv')
        grocery_data.rename(columns={'name': 'Items'}, inplace=True)

        models_folder = "Model/Price_Prediction/arima"
        with open(os.path.join(models_folder, "arima_models.pkl"), 'rb') as f:
            loaded_models = pickle.load(f)

        summary_data, plot_url = predict_and_save(selected_product, loaded_models, grocery_data)

    return render_template(
        'dashboard.html',
        user_id=user_id,
        products=products,
        search_query=search_query,
        selected_product=selected_product,
        summary_data=summary_data,
        plot_url=plot_url
    )


@app.route('/logout')
def logout():
    session.clear()  # Clear session
    flash('You have been logged out successfully!', 'info')
    return redirect(url_for('home'))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        user_id = request.form['user_id']
        password = request.form['password']

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                flash('User ID already exists!', 'error')
                return redirect(url_for('signup'))

        ws.append([user_id, password])
        wb.save(EXCEL_FILE)
        flash('Sign up successful! Please log in.', 'success')
        return redirect(url_for('home'))

    return render_template('signup.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        user_id = request.form['user_id']

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                flash(f'Password for {user_id} is: {row[1]}', 'info')
                return redirect(url_for('home'))

        flash('User ID not found!', 'error')
        return redirect(url_for('forgot_password'))

    return render_template('forgot_password.html')

@app.route('/info/<product_name>')
def info(product_name):
    # Example: Fetch product details
    product_details = {
        'Apple': 'Apples are rich in fiber and vitamins.',
        'Broccoli': 'Broccoli is a great source of vitamins K and C.',
        # Add more products as needed
    }
    info_text = product_details.get(product_name, "No information available.")
    return f"Information about {product_name}: {info_text}"

@app.route('/cart_add/<product_name>')
def cart_add(product_name):
    products = load_products()  # Load all products
    for product in products:
        if product['name'] == product_name:
            session['cart'].append(product)  # Add product to the session cart
            session.modified = True  # Notify Flask to update session
            flash(f'{product_name} has been added to your cart!', 'success')
            break
    else:
        flash(f'Product {product_name} not found.', 'error')
    return redirect(url_for('dashboard', user_id=session.get('user_id')))

@app.route('/cart')
def cart():
    return render_template('cart.html', cart=session['cart'])

@app.route('/cart_clear')
def cart_clear():
    session['cart'] = []  # Empty the cart
    session.modified = True
    flash('Your cart has been cleared.', 'info')
    return redirect(url_for('cart'))

if __name__ == '__main__':
    grocery_data = pd.read_csv('Datasets/Price_Predictions/grocery_items_bavaria.csv')
    grocery_data.rename(columns={'name': 'Items'}, inplace=True)

    csv_file_path = "Datasets/Output/Price_Prediction/Item_lists.csv"
    if os.path.exists(csv_file_path):
        os.remove(csv_file_path)
        print(f"Existing file {csv_file_path} has been deleted.")

    models_folder = "Model/Price_Prediction/arima"
    with open(os.path.join(models_folder, "arima_models.pkl"), 'rb') as f:
        loaded_models = pickle.load(f)

    items_to_predict = [
        "Allgäuer Hof-Milch Butter mild gesäuert 250g",
        "Bio Aubergine 1 Stück",
        "Blumenkohl weiß 1 Stück",
        "Broccoli 500g",
        "Eisbergsalat 1 Stück",
        "Galiamelone 1 Stück",
        "Karotten 1kg",
        "Kartoffeln vorwiegend festkochend 2,5kg",
        "Mango vorgereift 1 Stück",
        "Meggle Feine Butter 250g",
        "Orangen 2kg im Netz",
        "REWE Beste Wahl Feinschmecker Hähnchen 1200g",
        "REWE Bio Zucchini 500g",
        "Rewe Beste Wahl Eier aus Freilandhaltung 10 Stück",
        "Rispentomaten ca. 100g",
        "Spitzkohl ca. 1kg",
        "Tafeltrauben hell kernlos 500g",
        "Zitronen 500g im Netz",
        "Zwiebeln 2kg im Netz",
        "ja! Basmati Reis 1kg",
        "ja! H-Milch 3,5% 1",
        "ja! Sonnenblumenöl 1l"
    ]

    for item_name in items_to_predict:
        predict_and_save(item_name, loaded_models, grocery_data)

    initialize_excel()
    app.run(debug=True)
